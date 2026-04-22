#!/usr/bin/env python3
"""Build calibration_spreadsheet.xlsx, a blank template for 5 calibration events.

Jennifer fills in event_id/lesson/timestamp/event_text/context manually by
pulling from the US3 v10 transcript. Anna and Uyi rate; a third rater column
block (JK) holds Jennifer's pre-rated reference ratings revealed during the
norming meeting.
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUT_PATH = HERE / "calibration_spreadsheet.xlsx"

META_COLS = ["event_id", "lesson", "timestamp", "attribution", "speaker_label", "event_text", "context"]
RATER_COLS = ["factual", "temporal", "relevance", "detail", "notes"]
RATERS = [
    ("AB", "DCE7F1"),   # Anna, light blue
    ("UU", "D5E8D4"),   # Uyi, light green
    ("JK", "FFF2CC"),   # Jennifer reference, light yellow
]

HEADER_FILL = PatternFill("solid", fgColor="4A4A4A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
N_CALIBRATION_EVENTS = 5


def main() -> None:
    headers = list(META_COLS)
    fills: dict[int, str] = {}
    for rater, color in RATERS:
        for field in RATER_COLS:
            headers.append(f"{rater}_{field}")
            fills[len(headers)] = color

    wb = Workbook()
    ws = wb.active
    ws.title = "Calibration"

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, 2 + N_CALIBRATION_EVENTS):
        for col_idx in range(1, len(META_COLS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, color in fills.items():
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor=color)

    widths = {"event_id": 14, "lesson": 8, "timestamp": 10, "attribution": 12,
              "speaker_label": 16, "event_text": 40, "context": 50}
    for idx, name in enumerate(headers, start=1):
        if name in widths:
            w = widths[name]
        elif name.endswith("_notes"):
            w = 30
        else:
            w = 11
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({N_CALIBRATION_EVENTS} blank event rows, {len(headers)} columns)")


if __name__ == "__main__":
    main()
