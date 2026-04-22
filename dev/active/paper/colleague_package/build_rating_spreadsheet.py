#!/usr/bin/env python3
"""Build 04_rating_spreadsheet.xlsx from visual_validation_sample.csv.

Fans out the single rater-column block into two (AB + UU), adds header styling,
freezes panes, sets column widths, and applies fill colors per rater block.
"""
from __future__ import annotations
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
CSV_PATH = HERE.parent / "visual_validation_sample.csv"
OUT_PATH = HERE / "04_rating_spreadsheet.xlsx"

META_COLS = ["event_id", "lesson", "timestamp", "attribution", "speaker_label", "event_text", "context"]
RATER_COLS = ["factual", "temporal", "relevance", "detail", "notes"]
RATERS = [
    ("AB", "DCE7F1"),   # Anna Bloodworth, light blue
    ("UU", "D5E8D4"),   # Uyi Ugiagbe, light green
]

HEADER_FILL = PatternFill("solid", fgColor="4A4A4A")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def build_headers() -> tuple[list[str], dict[int, str]]:
    """Return (headers, col_index_to_fill_color) with 1-indexed column positions."""
    headers = list(META_COLS)
    fills: dict[int, str] = {}
    for rater, color in RATERS:
        for field in RATER_COLS:
            headers.append(f"{rater}_{field}")
            fills[len(headers)] = color
    return headers, fills


def main() -> None:
    if not CSV_PATH.exists():
        raise SystemExit(f"Input CSV missing: {CSV_PATH}")

    with CSV_PATH.open() as f:
        rows = list(csv.DictReader(f))
    if len(rows) != 45:
        print(f"WARNING: expected 45 rows, found {len(rows)}")

    headers, fills = build_headers()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ratings"

    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, r in enumerate(rows, start=2):
        for col_idx, field in enumerate(META_COLS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(field, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, color in fills.items():
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor=color)

    widths = {"event_id": 12, "lesson": 8, "timestamp": 10, "attribution": 12,
              "speaker_label": 16, "event_text": 40, "context": 50}
    for idx, name in enumerate(headers, start=1):
        if name in widths:
            w = widths[name]
        elif name.endswith("_notes"):
            w = 30
        else:
            w = 11
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({len(rows)} events, {len(headers)} columns)")


if __name__ == "__main__":
    main()
