import pandas as pd
import openpyxl

# Check source files
annaB_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_annaB.xlsx'
uyi_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_Uyi.xlsx'

print("=== ANNA B FILE (Full_Jake, Row 21) ===\n")
wb_anna = openpyxl.load_workbook(annaB_path)
ws_anna = wb_anna['Full_Jake']
for col_idx in range(1, 6):
    cell = ws_anna.cell(row=21, column=col_idx)
    print(f"Col {openpyxl.utils.get_column_letter(col_idx)}: {cell.value}")
wb_anna.close()

print("\n=== UYI FILE (Full_Jake, Row 21) ===\n")
wb_uyi = openpyxl.load_workbook(uyi_path)
ws_uyi = wb_uyi['Full_Jake']
for col_idx in range(1, 6):
    cell = ws_uyi.cell(row=21, column=col_idx)
    print(f"Col {openpyxl.utils.get_column_letter(col_idx)}: {cell.value}")
wb_uyi.close()

print("\n=== COMBINED FILE (Full_Jake, Row 21) ===\n")
combined_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_combined.xlsx'
wb_comb = openpyxl.load_workbook(combined_path)
ws_comb = wb_comb['Full_Jake']
for col_idx in range(1, 13):
    cell = ws_comb.cell(row=21, column=col_idx)
    print(f"Col {openpyxl.utils.get_column_letter(col_idx):2s}: {cell.value}")
wb_comb.close()

# Check another row with actual data
print("\n\n=== ROW 22 (First actual data row) ===\n")

wb_anna = openpyxl.load_workbook(annaB_path)
ws_anna = wb_anna['Full_Jake']
print("AnnaB:")
for col_idx in range(1, 6):
    cell = ws_anna.cell(row=22, column=col_idx)
    val = str(cell.value)[:50] if cell.value else "[empty]"
    print(f"  Col {openpyxl.utils.get_column_letter(col_idx)}: {val}")
wb_anna.close()

wb_uyi = openpyxl.load_workbook(uyi_path)
ws_uyi = wb_uyi['Full_Jake']
print("\nUyi:")
for col_idx in range(1, 6):
    cell = ws_uyi.cell(row=22, column=col_idx)
    val = str(cell.value)[:50] if cell.value else "[empty]"
    print(f"  Col {openpyxl.utils.get_column_letter(col_idx)}: {val}")
wb_uyi.close()

wb_comb = openpyxl.load_workbook(combined_path)
ws_comb = wb_comb['Full_Jake']
print("\nCombined:")
for col_idx in range(1, 13):
    cell = ws_comb.cell(row=22, column=col_idx)
    val = str(cell.value)[:50] if cell.value else "[empty]"
    print(f"  Col {openpyxl.utils.get_column_letter(col_idx):2s}: {val}")
wb_comb.close()
