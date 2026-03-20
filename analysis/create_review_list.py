import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

consensus_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_with_consensus.xlsx'

print("=== CREATING REVIEW LIST ===\n")

# Load workbook
wb = openpyxl.load_workbook(consensus_path)

# Create new sheet for review list
if 'REVIEW_LIST' in wb.sheetnames:
    del wb['REVIEW_LIST']

review_sheet = wb.create_sheet('REVIEW_LIST', 0)  # Insert as first sheet

# Set up headers
headers = [
    'Sheet', 'Row', 'Transcript',
    'Jen Speaker', 'Jen Utterance', 'Jen Notes',
    'Anna Speaker', 'Anna Utterance', 'Anna Notes',
    'Uyi Speaker', 'Uyi Utterance', 'Uyi Notes',
    'Speaker Agreement', 'Utterance Agreement',
    'FINAL DECISION: Speaker Error?', 'FINAL DECISION: Utterance Error?', 'FINAL DECISION: Notes'
]

for col_idx, header in enumerate(headers, 1):
    cell = review_sheet.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

# Collect all items needing review
review_items = []

for sheet_name in wb.sheetnames:
    if sheet_name == 'REVIEW_LIST':
        continue

    ws = wb[sheet_name]

    for row_idx in range(22, ws.max_row + 1):
        # Check if review needed
        review_needed = ws.cell(row=row_idx, column=19).value  # Column S
        if review_needed != "YES":
            continue

        # Get all data
        transcript = ws.cell(row=row_idx, column=1).value
        if not transcript:
            continue

        # Shorten transcript for readability
        transcript_short = str(transcript)[:100] + "..." if len(str(transcript)) > 100 else str(transcript)

        # Jennifer (C-E)
        jen_speaker = ws.cell(row=row_idx, column=3).value
        jen_utterance = ws.cell(row=row_idx, column=4).value
        jen_notes = ws.cell(row=row_idx, column=5).value

        # Anna (G-I)
        anna_speaker = ws.cell(row=row_idx, column=7).value
        anna_utterance = ws.cell(row=row_idx, column=8).value
        anna_notes = ws.cell(row=row_idx, column=9).value

        # Uyi (K-M)
        uyi_speaker = ws.cell(row=row_idx, column=11).value
        uyi_utterance = ws.cell(row=row_idx, column=12).value
        uyi_notes = ws.cell(row=row_idx, column=13).value

        # Agreement levels
        speaker_agreement = ws.cell(row=row_idx, column=17).value
        utterance_agreement = ws.cell(row=row_idx, column=18).value

        review_items.append({
            'sheet': sheet_name,
            'row': row_idx,
            'transcript': transcript_short,
            'jen_speaker': jen_speaker,
            'jen_utterance': jen_utterance,
            'jen_notes': jen_notes,
            'anna_speaker': anna_speaker,
            'anna_utterance': anna_utterance,
            'anna_notes': anna_notes,
            'uyi_speaker': uyi_speaker,
            'uyi_utterance': uyi_utterance,
            'uyi_notes': uyi_notes,
            'speaker_agreement': speaker_agreement,
            'utterance_agreement': utterance_agreement
        })

print(f"Found {len(review_items)} items needing review\n")

# Write to review sheet
for idx, item in enumerate(review_items, 2):  # Start at row 2
    review_sheet.cell(row=idx, column=1).value = item['sheet']
    review_sheet.cell(row=idx, column=2).value = item['row']
    review_sheet.cell(row=idx, column=3).value = item['transcript']

    review_sheet.cell(row=idx, column=4).value = item['jen_speaker']
    review_sheet.cell(row=idx, column=5).value = item['jen_utterance']
    review_sheet.cell(row=idx, column=6).value = item['jen_notes']

    review_sheet.cell(row=idx, column=7).value = item['anna_speaker']
    review_sheet.cell(row=idx, column=8).value = item['anna_utterance']
    review_sheet.cell(row=idx, column=9).value = item['anna_notes']

    review_sheet.cell(row=idx, column=10).value = item['uyi_speaker']
    review_sheet.cell(row=idx, column=11).value = item['uyi_utterance']
    review_sheet.cell(row=idx, column=12).value = item['uyi_notes']

    review_sheet.cell(row=idx, column=13).value = item['speaker_agreement']
    review_sheet.cell(row=idx, column=14).value = item['utterance_agreement']

    # Final decision columns (empty for user to fill)
    review_sheet.cell(row=idx, column=15).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    review_sheet.cell(row=idx, column=16).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    review_sheet.cell(row=idx, column=17).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Set column widths
review_sheet.column_dimensions['A'].width = 15  # Sheet
review_sheet.column_dimensions['B'].width = 8   # Row
review_sheet.column_dimensions['C'].width = 60  # Transcript
review_sheet.column_dimensions['D'].width = 12  # Jen Speaker
review_sheet.column_dimensions['E'].width = 12  # Jen Utterance
review_sheet.column_dimensions['F'].width = 40  # Jen Notes
review_sheet.column_dimensions['G'].width = 12  # Anna Speaker
review_sheet.column_dimensions['H'].width = 12  # Anna Utterance
review_sheet.column_dimensions['I'].width = 40  # Anna Notes
review_sheet.column_dimensions['J'].width = 12  # Uyi Speaker
review_sheet.column_dimensions['K'].width = 12  # Uyi Utterance
review_sheet.column_dimensions['L'].width = 40  # Uyi Notes
review_sheet.column_dimensions['M'].width = 18  # Speaker Agreement
review_sheet.column_dimensions['N'].width = 18  # Utterance Agreement
review_sheet.column_dimensions['O'].width = 25  # Final: Speaker
review_sheet.column_dimensions['P'].width = 25  # Final: Utterance
review_sheet.column_dimensions['Q'].width = 50  # Final: Notes

# Freeze panes
review_sheet.freeze_panes = 'D2'  # Freeze first 3 columns and header row

# Save
wb.save(consensus_path)
wb.close()

print(f"✅ Review list created!")
print(f"   Total items: {len(review_items)}")
print(f"   Location: REVIEW_LIST sheet (first tab)")
print(f"\nInstructions:")
print(f"   1. Open transcripts_with_consensus.xlsx")
print(f"   2. Go to REVIEW_LIST tab")
print(f"   3. Fill in yellow columns (O, P, Q) with your final decisions")
print(f"   4. Format: YES/NO for speaker/utterance errors")
print(f"   5. Add corrected speaker/text in Notes column")
