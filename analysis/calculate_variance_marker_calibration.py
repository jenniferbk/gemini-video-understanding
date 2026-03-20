import openpyxl
import re
from collections import defaultdict

consensus_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_with_consensus.xlsx'

print("=== VARIANCE MARKER CALIBRATION ANALYSIS ===\n")
print("Analyzing how well variance markers (⚠️, 🚨, [uncertain]) predict errors\n")

# Load workbook
wb = openpyxl.load_workbook(consensus_path)

# Statistics
stats = {
    'total_utterances': 0,

    # Speaker ID
    'speaker_error_with_marker': 0,      # True positive
    'speaker_error_no_marker': 0,        # False negative
    'speaker_no_error_with_marker': 0,   # False positive
    'speaker_no_error_no_marker': 0,     # True negative

    # Utterance
    'utterance_error_with_marker': 0,
    'utterance_error_no_marker': 0,
    'utterance_no_error_with_marker': 0,
    'utterance_no_error_no_marker': 0,

    # By condition
    'by_condition': defaultdict(lambda: {
        'total': 0,
        'speaker_error_with_marker': 0,
        'speaker_error_no_marker': 0,
        'speaker_no_error_with_marker': 0,
        'speaker_no_error_no_marker': 0
    })
}

for sheet_name in wb.sheetnames:
    if sheet_name in ['REVIEW_LIST', 'RESEARCH_SUMMARY']:
        continue

    # Determine condition
    if sheet_name.startswith('Full_'):
        condition = 'Full Class'
    elif sheet_name.startswith('SG_'):
        condition = 'Small Group'
    elif sheet_name.startswith('FG_'):
        condition = 'Focus Group'
    elif sheet_name.startswith('Post_'):
        condition = 'Post-Interview'
    else:
        continue

    ws = wb[sheet_name]

    for row_idx in range(22, ws.max_row + 1):
        # Skip if no transcript
        transcript = ws.cell(row=row_idx, column=1).value
        if not transcript or str(transcript).strip() == "":
            continue

        stats['total_utterances'] += 1
        stats['by_condition'][condition]['total'] += 1

        # Check for variance markers in transcript
        transcript_str = str(transcript)
        has_marker = bool('[uncertain]' in transcript_str or
                         '⚠️' in transcript_str or
                         '🚨' in transcript_str)

        # Get consensus decisions (worst case: any flagging counts as error)
        consensus_speaker = ws.cell(row=row_idx, column=14).value  # Col N
        consensus_utterance = ws.cell(row=row_idx, column=15).value  # Col O

        # Speaker ID analysis
        is_speaker_error = consensus_speaker and (
            consensus_speaker == "YES - Error" or consensus_speaker == "REVIEW"
        )

        if is_speaker_error and has_marker:
            stats['speaker_error_with_marker'] += 1
            stats['by_condition'][condition]['speaker_error_with_marker'] += 1
        elif is_speaker_error and not has_marker:
            stats['speaker_error_no_marker'] += 1
            stats['by_condition'][condition]['speaker_error_no_marker'] += 1
        elif not is_speaker_error and has_marker:
            stats['speaker_no_error_with_marker'] += 1
            stats['by_condition'][condition]['speaker_no_error_with_marker'] += 1
        else:  # not is_speaker_error and not has_marker
            stats['speaker_no_error_no_marker'] += 1
            stats['by_condition'][condition]['speaker_no_error_no_marker'] += 1

        # Utterance analysis
        is_utterance_error = consensus_utterance and (
            consensus_utterance == "YES - Error" or consensus_utterance == "REVIEW"
        )

        if is_utterance_error and has_marker:
            stats['utterance_error_with_marker'] += 1
        elif is_utterance_error and not has_marker:
            stats['utterance_error_no_marker'] += 1
        elif not is_utterance_error and has_marker:
            stats['utterance_no_error_with_marker'] += 1
        else:
            stats['utterance_no_error_no_marker'] += 1

wb.close()

# Calculate metrics
def calculate_metrics(tp, fp, fn, tn):
    """Calculate classification metrics"""
    total = tp + fp + fn + tn

    # Precision: Of marked items, what % are actual errors?
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0

    # Recall (Sensitivity): Of actual errors, what % are marked?
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0

    # Specificity: Of non-errors, what % are correctly unmarked?
    specificity = tn / (tn + fp) if (tn + fp) > 0 else 0

    # False Positive Rate: Of non-errors, what % are incorrectly marked?
    fpr = fp / (fp + tn) if (fp + tn) > 0 else 0

    # Accuracy: Overall correct classifications
    accuracy = (tp + tn) / total if total > 0 else 0

    # F1 Score: Harmonic mean of precision and recall
    f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0

    return {
        'precision': precision,
        'recall': recall,
        'specificity': specificity,
        'fpr': fpr,
        'accuracy': accuracy,
        'f1': f1
    }

# Speaker ID metrics
speaker_metrics = calculate_metrics(
    stats['speaker_error_with_marker'],      # TP
    stats['speaker_no_error_with_marker'],   # FP
    stats['speaker_error_no_marker'],        # FN
    stats['speaker_no_error_no_marker']      # TN
)

# Utterance metrics
utterance_metrics = calculate_metrics(
    stats['utterance_error_with_marker'],
    stats['utterance_no_error_with_marker'],
    stats['utterance_error_no_marker'],
    stats['utterance_no_error_no_marker']
)

# Print results
print("="*80)
print("SPEAKER ID - VARIANCE MARKER PERFORMANCE")
print("="*80)
print("\nConfusion Matrix:")
print(f"{'':>30} {'Actual Error':>20} {'Actual No Error':>20}")
print(f"{'Has Variance Marker':>30} {stats['speaker_error_with_marker']:>20} {stats['speaker_no_error_with_marker']:>20}")
print(f"{'No Variance Marker':>30} {stats['speaker_error_no_marker']:>20} {stats['speaker_no_error_no_marker']:>20}")

print(f"\n{'Total with markers':>30} {stats['speaker_error_with_marker'] + stats['speaker_no_error_with_marker']:>20}")
print(f"{'Total without markers':>30} {stats['speaker_error_no_marker'] + stats['speaker_no_error_no_marker']:>20}")
print(f"{'Total errors':>30} {stats['speaker_error_with_marker'] + stats['speaker_error_no_marker']:>20}")
print(f"{'Total no errors':>30} {stats['speaker_no_error_with_marker'] + stats['speaker_no_error_no_marker']:>20}")

print("\n" + "="*80)
print("CLASSIFICATION METRICS - SPEAKER ID")
print("="*80)
print(f"\nPrecision:     {speaker_metrics['precision']*100:>6.2f}%  (Of marked utterances, what % are actual errors?)")
print(f"Recall:        {speaker_metrics['recall']*100:>6.2f}%  (Of actual errors, what % are marked?)")
print(f"Specificity:   {speaker_metrics['specificity']*100:>6.2f}%  (Of non-errors, what % are correctly unmarked?)")
print(f"FPR:           {speaker_metrics['fpr']*100:>6.2f}%  (Of non-errors, what % are falsely marked?)")
print(f"Accuracy:      {speaker_metrics['accuracy']*100:>6.2f}%  (Overall correct classifications)")
print(f"F1 Score:      {speaker_metrics['f1']:>6.3f}   (Harmonic mean of precision/recall)")

print("\n" + "="*80)
print("UTTERANCE - VARIANCE MARKER PERFORMANCE")
print("="*80)
print("\nConfusion Matrix:")
print(f"{'':>30} {'Actual Error':>20} {'Actual No Error':>20}")
print(f"{'Has Variance Marker':>30} {stats['utterance_error_with_marker']:>20} {stats['utterance_no_error_with_marker']:>20}")
print(f"{'No Variance Marker':>30} {stats['utterance_error_no_marker']:>20} {stats['utterance_no_error_no_marker']:>20}")

print("\n" + "="*80)
print("CLASSIFICATION METRICS - UTTERANCE")
print("="*80)
print(f"\nPrecision:     {utterance_metrics['precision']*100:>6.2f}%  (Of marked utterances, what % are actual errors?)")
print(f"Recall:        {utterance_metrics['recall']*100:>6.2f}%  (Of actual errors, what % are marked?)")
print(f"Specificity:   {utterance_metrics['specificity']*100:>6.2f}%  (Of non-errors, what % are correctly unmarked?)")
print(f"FPR:           {utterance_metrics['fpr']*100:>6.2f}%  (Of non-errors, what % are falsely marked?)")
print(f"Accuracy:      {utterance_metrics['accuracy']*100:>6.2f}%  (Overall correct classifications)")
print(f"F1 Score:      {utterance_metrics['f1']:>6.3f}   (Harmonic mean of precision/recall)")

# By condition
print("\n\n" + "="*80)
print("SPEAKER ID VARIANCE MARKER PERFORMANCE BY CONDITION")
print("="*80)

for condition in ['Full Class', 'Small Group', 'Focus Group', 'Post-Interview']:
    cond_stats = stats['by_condition'][condition]
    if cond_stats['total'] == 0:
        continue

    metrics = calculate_metrics(
        cond_stats['speaker_error_with_marker'],
        cond_stats['speaker_no_error_with_marker'],
        cond_stats['speaker_error_no_marker'],
        cond_stats['speaker_no_error_no_marker']
    )

    total_markers = cond_stats['speaker_error_with_marker'] + cond_stats['speaker_no_error_with_marker']

    print(f"\n{condition.upper()}")
    print("-" * 70)
    print(f"Total utterances: {cond_stats['total']}")
    print(f"Total with markers: {total_markers}")
    print(f"  • True positives (error + marker): {cond_stats['speaker_error_with_marker']}")
    print(f"  • False positives (no error + marker): {cond_stats['speaker_no_error_with_marker']}")
    print(f"\nMetrics:")
    print(f"  Precision:    {metrics['precision']*100:>6.2f}%")
    print(f"  Recall:       {metrics['recall']*100:>6.2f}%")
    print(f"  FPR:          {metrics['fpr']*100:>6.2f}%")
    print(f"  F1 Score:     {metrics['f1']:>6.3f}")

# Summary interpretation
print("\n\n" + "="*80)
print("KEY FINDINGS - VARIANCE MARKER CALIBRATION")
print("="*80)
print(f"""
SPEAKER ID MARKERS:
• Precision: {speaker_metrics['precision']*100:.1f}% - When a marker is present, there's a {speaker_metrics['precision']*100:.1f}% chance it's an actual error
• Recall: {speaker_metrics['recall']*100:.1f}% - Catches {speaker_metrics['recall']*100:.1f}% of all actual errors
• False Positive Rate: {speaker_metrics['fpr']*100:.1f}% - {speaker_metrics['fpr']*100:.1f}% of correct transcriptions are falsely flagged
• Specificity: {speaker_metrics['specificity']*100:.1f}% - {speaker_metrics['specificity']*100:.1f}% of correct transcriptions correctly have no marker

UTTERANCE MARKERS:
• Precision: {utterance_metrics['precision']*100:.1f}%
• Recall: {utterance_metrics['recall']*100:.1f}%
• False Positive Rate: {utterance_metrics['fpr']*100:.1f}%

INTERPRETATION:
The variance markers are {'well-calibrated' if speaker_metrics['precision'] > 0.5 else 'poorly calibrated'} for speaker ID errors.
{'✅ Good precision: Most marked items are actually errors' if speaker_metrics['precision'] > 0.5 else '⚠️ Low precision: Many marked items are not errors'}
{'✅ Good recall: Catches most errors' if speaker_metrics['recall'] > 0.7 else '⚠️ Moderate recall: Misses some errors'}
{'✅ Low false positive rate: Few false alarms' if speaker_metrics['fpr'] < 0.2 else '⚠️ High false positive rate: Many false alarms'}

PRACTICAL IMPLICATIONS:
{f"• If you review all {stats['speaker_error_with_marker'] + stats['speaker_no_error_with_marker']} marked utterances, you'll find {stats['speaker_error_with_marker']} actual errors" if (stats['speaker_error_with_marker'] + stats['speaker_no_error_with_marker']) > 0 else "• No marked utterances to review"}
{f"• That's {stats['speaker_no_error_with_marker']} false positives (wasted reviews)" if stats['speaker_no_error_with_marker'] > 0 else ""}
{f"• You'll miss {stats['speaker_error_no_marker']} errors that weren't marked" if stats['speaker_error_no_marker'] > 0 else ""}
• Still better than reviewing all {stats['total_utterances']} utterances!
""")

print("\n✅ Variance marker calibration analysis complete!")
