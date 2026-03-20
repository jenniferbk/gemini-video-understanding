import openpyxl

jen_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_jen.xlsx'
output_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_all_combined.xlsx'

print("=== CHECKING JENNIFER'S SOURCE DATA ===\n")

wb_jen = openpyxl.load_workbook(jen_path)
ws_jen = wb_jen['Full_Jake']

print("Jennifer's file - Full_Jake, rows 22-30 (cols C-E):")
for row_idx in range(22, 31):
    row_data = []
    for col_idx in [3, 4, 5]:  # C, D, E
        cell = ws_jen.cell(row=row_idx, column=col_idx)
        val = str(cell.value)[:40] if cell.value else "[empty]"
        row_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}={val}")
    print(f"  Row {row_idx}: {' | '.join(row_data)}")

wb_jen.close()

print("\n=== CHECKING COMBINED OUTPUT ===\n")

wb_out = openpyxl.load_workbook(output_path)
ws_out = wb_out['Full_Jake']

print("Combined file - Full_Jake, rows 22-30 (cols C-E for Jennifer):")
for row_idx in range(22, 31):
    row_data = []
    for col_idx in [3, 4, 5]:  # C, D, E
        cell = ws_out.cell(row=row_idx, column=col_idx)
        val = str(cell.value)[:40] if cell.value else "[empty]"
        row_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}={val}")
    print(f"  Row {row_idx}: {' | '.join(row_data)}")

wb_out.close()
