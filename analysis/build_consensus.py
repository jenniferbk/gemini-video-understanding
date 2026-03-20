import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import pandas as pd

combined_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_all_combined.xlsx'
output_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_with_consensus.xlsx'

print("=== BUILDING CONSENSUS ANALYSIS ===\n")

# Load workbook
wb = openpyxl.load_workbook(combined_path)

# Colors for highlighting
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 2/3 agreement
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")     # 1/3 or no agreement
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # 3/3 agreement

# Process each sheet
for sheet_name in wb.sheetnames:
    print(f"Processing sheet: {sheet_name}")
    ws = wb[sheet_name]

    # Add consensus column headers in row 21
    ws.cell(row=19, column=14).value = "CONSENSUS ANALYSIS"
    ws.cell(row=19, column=14).font = Font(bold=True, size=14)

    ws.cell(row=20, column=14).value = "Consensus"
    ws.cell(row=21, column=14).value = "Speaker Error?"

    ws.cell(row=20, column=15).value = "Consensus"
    ws.cell(row=21, column=15).value = "Utterance Error?"

    ws.cell(row=20, column=16).value = "Consensus"
    ws.cell(row=21, column=16).value = "Notes"

    ws.cell(row=21, column=17).value = "Speaker Agreement"
    ws.cell(row=21, column=18).value = "Utterance Agreement"
    ws.cell(row=21, column=19).value = "Review Needed?"

    # Make headers bold
    for col in range(14, 20):
        ws.cell(row=21, column=col).font = Font(bold=True)

    # Track statistics
    stats = {
        'total_rows': 0,
        'speaker_3_3': 0,
        'speaker_2_3': 0,
        'speaker_1_3': 0,
        'speaker_0_3': 0,
        'utterance_3_3': 0,
        'utterance_2_3': 0,
        'utterance_1_3': 0,
        'utterance_0_3': 0,
        'needs_review': 0
    }

    # Process data rows (starting from row 22)
    for row_idx in range(22, ws.max_row + 1):
        # Skip if no transcript
        transcript = ws.cell(row=row_idx, column=1).value
        if not transcript or str(transcript).strip() == "":
            continue

        stats['total_rows'] += 1

        # Get coding from each person
        # Jennifer (C-E): cols 3, 4, 5
        jen_speaker = ws.cell(row=row_idx, column=3).value
        jen_utterance = ws.cell(row=row_idx, column=4).value
        jen_notes = ws.cell(row=row_idx, column=5).value

        # Anna (G-I): cols 7, 8, 9
        anna_speaker = ws.cell(row=row_idx, column=7).value
        anna_utterance = ws.cell(row=row_idx, column=8).value
        anna_notes = ws.cell(row=row_idx, column=9).value

        # Uyi (K-M): cols 11, 12, 13
        uyi_speaker = ws.cell(row=row_idx, column=11).value
        uyi_utterance = ws.cell(row=row_idx, column=12).value
        uyi_notes = ws.cell(row=row_idx, column=13).value

        # Analyze Speaker ID coding
        # Count how many flagged this as having a speaker error (any non-empty value)
        speaker_flags = []
        if jen_speaker and str(jen_speaker).strip():
            speaker_flags.append('Jennifer')
        if anna_speaker and str(anna_speaker).strip():
            speaker_flags.append('Anna')
        if uyi_speaker and str(uyi_speaker).strip():
            speaker_flags.append('Uyi')

        speaker_count = len(speaker_flags)

        # Analyze Utterance coding
        utterance_flags = []
        if jen_utterance and str(jen_utterance).strip():
            utterance_flags.append('Jennifer')
        if anna_utterance and str(anna_utterance).strip():
            utterance_flags.append('Anna')
        if uyi_utterance and str(uyi_utterance).strip():
            utterance_flags.append('Uyi')

        utterance_count = len(utterance_flags)

        # Determine consensus for Speaker ID
        consensus_speaker = ""
        speaker_agreement = ""
        if speaker_count == 3:
            consensus_speaker = "YES - Error"
            speaker_agreement = "3/3"
            stats['speaker_3_3'] += 1
            ws.cell(row=row_idx, column=14).fill = green_fill
        elif speaker_count == 2:
            consensus_speaker = "REVIEW"
            speaker_agreement = f"2/3 ({', '.join(speaker_flags)})"
            stats['speaker_2_3'] += 1
            ws.cell(row=row_idx, column=14).fill = yellow_fill
        elif speaker_count == 1:
            consensus_speaker = "REVIEW"
            speaker_agreement = f"1/3 ({speaker_flags[0]})"
            stats['speaker_1_3'] += 1
            ws.cell(row=row_idx, column=14).fill = red_fill
        else:
            consensus_speaker = "No error"
            speaker_agreement = "0/3"
            stats['speaker_0_3'] += 1

        # Determine consensus for Utterance
        consensus_utterance = ""
        utterance_agreement = ""
        if utterance_count == 3:
            consensus_utterance = "YES - Error"
            utterance_agreement = "3/3"
            stats['utterance_3_3'] += 1
            ws.cell(row=row_idx, column=15).fill = green_fill
        elif utterance_count == 2:
            consensus_utterance = "REVIEW"
            utterance_agreement = f"2/3 ({', '.join(utterance_flags)})"
            stats['utterance_2_3'] += 1
            ws.cell(row=row_idx, column=15).fill = yellow_fill
        elif utterance_count == 1:
            consensus_utterance = "REVIEW"
            utterance_agreement = f"1/3 ({utterance_flags[0]})"
            stats['utterance_1_3'] += 1
            ws.cell(row=row_idx, column=15).fill = red_fill
        else:
            consensus_utterance = "No error"
            utterance_agreement = "0/3"
            stats['utterance_0_3'] += 1

        # Combine notes where there's agreement
        consensus_notes = ""
        if speaker_count == 3 or utterance_count == 3:
            # Combine all notes
            all_notes = []
            if jen_notes:
                all_notes.append(f"Jen: {jen_notes}")
            if anna_notes:
                all_notes.append(f"Anna: {anna_notes}")
            if uyi_notes:
                all_notes.append(f"Uyi: {uyi_notes}")
            consensus_notes = " | ".join(all_notes)

        # Review needed?
        review_needed = ""
        if "REVIEW" in consensus_speaker or "REVIEW" in consensus_utterance:
            review_needed = "YES"
            stats['needs_review'] += 1
            ws.cell(row=row_idx, column=19).fill = yellow_fill
            ws.cell(row=row_idx, column=19).font = Font(bold=True)

        # Write consensus columns
        ws.cell(row=row_idx, column=14).value = consensus_speaker
        ws.cell(row=row_idx, column=15).value = consensus_utterance
        ws.cell(row=row_idx, column=16).value = consensus_notes
        ws.cell(row=row_idx, column=17).value = speaker_agreement
        ws.cell(row=row_idx, column=18).value = utterance_agreement
        ws.cell(row=row_idx, column=19).value = review_needed

    # Print statistics for this sheet
    print(f"  Statistics for {sheet_name}:")
    print(f"    Total transcript rows: {stats['total_rows']}")
    print(f"    Speaker ID - 3/3 agreement: {stats['speaker_3_3']}")
    print(f"    Speaker ID - 2/3 agreement: {stats['speaker_2_3']}")
    print(f"    Speaker ID - 1/3 agreement: {stats['speaker_1_3']}")
    print(f"    Utterance - 3/3 agreement: {stats['utterance_3_3']}")
    print(f"    Utterance - 2/3 agreement: {stats['utterance_2_3']}")
    print(f"    Utterance - 1/3 agreement: {stats['utterance_1_3']}")
    print(f"    Rows needing review: {stats['needs_review']}")
    print()

# Save with consensus
print(f"Saving to: {output_path}")
wb.save(output_path)
wb.close()

print("\n✅ Consensus analysis complete!")
print("\nColor coding:")
print("  🟢 Green = 3/3 agreement (auto-accepted)")
print("  🟡 Yellow = 2/3 agreement (needs review)")
print("  🔴 Red = 1/3 or less agreement (needs review)")
print("\nNew columns added (N-S):")
print("  N: Consensus Speaker Error")
print("  O: Consensus Utterance Error")
print("  P: Consensus Notes")
print("  Q: Speaker Agreement level")
print("  R: Utterance Agreement level")
print("  S: Review Needed flag")
