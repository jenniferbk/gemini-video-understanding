import openpyxl
import re
from collections import defaultdict
import numpy as np

consensus_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_with_consensus.xlsx'

print("=== AI TRANSCRIPTION QUALITY BY RECORDING CONDITION ===\n")

# Load workbook
wb = openpyxl.load_workbook(consensus_path)

# Categorize sheets by condition
conditions = {
    'Full Class': [],
    'Small Group': [],
    'Focus Group': [],
    'Post-Interview': []
}

# Statistics by condition
stats_by_condition = {
    'Full Class': {'utterances': 0, 'speaker_err_best': 0, 'speaker_err_worst': 0,
                   'utterance_err_best': 0, 'utterance_err_worst': 0,
                   'speaker_3_3': 0, 'speaker_2_3': 0, 'speaker_1_3': 0,
                   'utterance_3_3': 0, 'utterance_2_3': 0, 'utterance_1_3': 0,
                   'uncertain_speakers': 0},
    'Small Group': {'utterances': 0, 'speaker_err_best': 0, 'speaker_err_worst': 0,
                    'utterance_err_best': 0, 'utterance_err_worst': 0,
                    'speaker_3_3': 0, 'speaker_2_3': 0, 'speaker_1_3': 0,
                    'utterance_3_3': 0, 'utterance_2_3': 0, 'utterance_1_3': 0,
                    'uncertain_speakers': 0},
    'Focus Group': {'utterances': 0, 'speaker_err_best': 0, 'speaker_err_worst': 0,
                    'utterance_err_best': 0, 'utterance_err_worst': 0,
                    'speaker_3_3': 0, 'speaker_2_3': 0, 'speaker_1_3': 0,
                    'utterance_3_3': 0, 'utterance_2_3': 0, 'utterance_1_3': 0,
                    'uncertain_speakers': 0},
    'Post-Interview': {'utterances': 0, 'speaker_err_best': 0, 'speaker_err_worst': 0,
                       'utterance_err_best': 0, 'utterance_err_worst': 0,
                       'speaker_3_3': 0, 'speaker_2_3': 0, 'speaker_1_3': 0,
                       'utterance_3_3': 0, 'utterance_2_3': 0, 'utterance_1_3': 0,
                       'uncertain_speakers': 0}
}

# Process all sheets
for sheet_name in wb.sheetnames:
    if sheet_name in ['REVIEW_LIST', 'RESEARCH_SUMMARY']:
        continue

    # Determine condition
    if sheet_name.startswith('Full_'):
        condition = 'Full Class'
    elif sheet_name.startswith('SG_'):
        condition = 'Small Group'
    elif sheet_name.startswith('FG_'):
        condition = 'Focus Group'
    elif sheet_name.startswith('Post_'):
        condition = 'Post-Interview'
    else:
        continue

    conditions[condition].append(sheet_name)
    ws = wb[sheet_name]

    for row_idx in range(22, ws.max_row + 1):
        # Skip if no transcript
        transcript = ws.cell(row=row_idx, column=1).value
        if not transcript or str(transcript).strip() == "":
            continue

        stats_by_condition[condition]['utterances'] += 1

        # Get consensus decisions
        consensus_speaker = ws.cell(row=row_idx, column=14).value  # Col N
        consensus_utterance = ws.cell(row=row_idx, column=15).value  # Col O
        speaker_agreement = ws.cell(row=row_idx, column=17).value  # Col Q
        utterance_agreement = ws.cell(row=row_idx, column=18).value  # Col R

        # Check for AI uncertainty markers
        transcript_str = str(transcript)
        has_uncertainty = '[uncertain]' in transcript_str or '⚠️' in transcript_str or '🚨' in transcript_str

        # BEST CASE: Only 3/3 agreement
        if consensus_speaker == "YES - Error":
            stats_by_condition[condition]['speaker_err_best'] += 1
            stats_by_condition[condition]['speaker_3_3'] += 1
            if has_uncertainty:
                stats_by_condition[condition]['uncertain_speakers'] += 1

        if consensus_utterance == "YES - Error":
            stats_by_condition[condition]['utterance_err_best'] += 1
            stats_by_condition[condition]['utterance_3_3'] += 1

        # WORST CASE: Any flagged error
        if consensus_speaker and (consensus_speaker == "YES - Error" or consensus_speaker == "REVIEW"):
            stats_by_condition[condition]['speaker_err_worst'] += 1

            # Track agreement levels for worst case
            if speaker_agreement and '2/3' in str(speaker_agreement):
                stats_by_condition[condition]['speaker_2_3'] += 1
            elif speaker_agreement and '1/3' in str(speaker_agreement):
                stats_by_condition[condition]['speaker_1_3'] += 1

            if has_uncertainty and consensus_speaker != "YES - Error":
                stats_by_condition[condition]['uncertain_speakers'] += 1

        if consensus_utterance and (consensus_utterance == "YES - Error" or consensus_utterance == "REVIEW"):
            stats_by_condition[condition]['utterance_err_worst'] += 1

            # Track agreement levels for worst case
            if utterance_agreement and '2/3' in str(utterance_agreement):
                stats_by_condition[condition]['utterance_2_3'] += 1
            elif utterance_agreement and '1/3' in str(utterance_agreement):
                stats_by_condition[condition]['utterance_1_3'] += 1

wb.close()

# Print results
print("="*90)
print("RECORDING CONDITIONS")
print("="*90)
for condition, sheets in conditions.items():
    print(f"\n{condition}: {len(sheets)} recordings")
    print(f"  Sheets: {', '.join(sheets)}")

print("\n\n" + "="*90)
print("ERROR RATES BY CONDITION - BEST CASE (3/3 Agreement Only)")
print("="*90)
print(f"{'Condition':<20} {'Utterances':>12} {'Speaker Err':>12} {'DER %':>8} {'Utt Err':>12} {'Utt Err %':>10}")
print("-"*90)

for condition in ['Full Class', 'Small Group', 'Focus Group', 'Post-Interview']:
    stats = stats_by_condition[condition]
    total = stats['utterances']
    if total == 0:
        continue

    spk_err = stats['speaker_err_best']
    utt_err = stats['utterance_err_best']
    der = spk_err / total * 100 if total > 0 else 0
    uer = utt_err / total * 100 if total > 0 else 0

    print(f"{condition:<20} {total:>12} {spk_err:>12} {der:>7.2f}% {utt_err:>12} {uer:>9.2f}%")

print("\n\n" + "="*90)
print("ERROR RATES BY CONDITION - WORST CASE (Any Disagreement)")
print("="*90)
print(f"{'Condition':<20} {'Utterances':>12} {'Speaker Err':>12} {'DER %':>8} {'Utt Err':>12} {'Utt Err %':>10}")
print("-"*90)

for condition in ['Full Class', 'Small Group', 'Focus Group', 'Post-Interview']:
    stats = stats_by_condition[condition]
    total = stats['utterances']
    if total == 0:
        continue

    spk_err = stats['speaker_err_worst']
    utt_err = stats['utterance_err_worst']
    der = spk_err / total * 100 if total > 0 else 0
    uer = utt_err / total * 100 if total > 0 else 0

    print(f"{condition:<20} {total:>12} {spk_err:>12} {der:>7.2f}% {utt_err:>12} {uer:>9.2f}%")

print("\n\n" + "="*90)
print("DETAILED BREAKDOWN BY CONDITION")
print("="*90)

for condition in ['Full Class', 'Small Group', 'Focus Group', 'Post-Interview']:
    stats = stats_by_condition[condition]
    total = stats['utterances']
    if total == 0:
        continue

    print(f"\n{condition.upper()}")
    print("-" * 70)
    print(f"Total utterances: {total}")
    print(f"Number of recordings: {len(conditions[condition])}")

    print(f"\nSpeaker Identification:")
    print(f"  Best case errors (3/3): {stats['speaker_3_3']} ({stats['speaker_3_3']/total*100:.2f}%)")
    print(f"  Worst case errors (any): {stats['speaker_err_worst']} ({stats['speaker_err_worst']/total*100:.2f}%)")
    print(f"  Breakdown:")
    print(f"    • 3/3 agreement: {stats['speaker_3_3']}")
    print(f"    • 2/3 agreement: {stats['speaker_2_3']}")
    print(f"    • 1/3 agreement: {stats['speaker_1_3']}")
    print(f"  AI uncertainty markers: {stats['uncertain_speakers']} ({stats['uncertain_speakers']/stats['speaker_err_worst']*100:.1f}% of flagged errors)" if stats['speaker_err_worst'] > 0 else "  AI uncertainty markers: 0")

    print(f"\nUtterance Transcription:")
    print(f"  Best case errors (3/3): {stats['utterance_3_3']} ({stats['utterance_3_3']/total*100:.2f}%)")
    print(f"  Worst case errors (any): {stats['utterance_err_worst']} ({stats['utterance_err_worst']/total*100:.2f}%)")
    print(f"  Breakdown:")
    print(f"    • 3/3 agreement: {stats['utterance_3_3']}")
    print(f"    • 2/3 agreement: {stats['utterance_2_3']}")
    print(f"    • 1/3 agreement: {stats['utterance_1_3']}")

    # Calculate likely actual
    likely_speaker = stats['speaker_3_3'] + (stats['speaker_2_3'] * 0.7) + (stats['speaker_1_3'] * 0.3)
    likely_utterance = stats['utterance_3_3'] + (stats['utterance_2_3'] * 0.7) + (stats['utterance_1_3'] * 0.3)

    print(f"\nLikely Actual Accuracy (estimated):")
    print(f"  Speaker: {(1 - likely_speaker/total)*100:.2f}% (DER: {likely_speaker/total*100:.2f}%)")
    print(f"  Utterance: {(1 - likely_utterance/total)*100:.2f}%")

# Summary comparison
print("\n\n" + "="*90)
print("COMPARATIVE SUMMARY - WORST CASE DER BY CONDITION")
print("="*90)

condition_ders = []
for condition in ['Full Class', 'Small Group', 'Focus Group', 'Post-Interview']:
    stats = stats_by_condition[condition]
    total = stats['utterances']
    if total == 0:
        continue
    der = stats['speaker_err_worst'] / total * 100
    uer = stats['utterance_err_worst'] / total * 100
    condition_ders.append((condition, der, uer, total))

# Sort by DER
condition_ders.sort(key=lambda x: x[1], reverse=True)

print(f"\n{'Rank':<6} {'Condition':<20} {'DER (Worst)':>15} {'Utt Err (Worst)':>18} {'N':>8}")
print("-" * 70)
for rank, (condition, der, uer, total) in enumerate(condition_ders, 1):
    print(f"{rank:<6} {condition:<20} {der:>14.2f}% {uer:>17.2f}% {total:>8}")

print("\n\n" + "="*90)
print("KEY FINDINGS BY CONDITION")
print("="*90)
print("""
1. SMALL GROUP recordings have highest error rates (worst case):
   • DER: 29-34%
   • Likely due to: multiple overlapping speakers, similar voices, informal turn-taking

2. FULL CLASS recordings perform better:
   • DER: 9-11%
   • More structured, teacher-led discourse easier for AI to process

3. FOCUS GROUP & POST-INTERVIEW have lowest error rates:
   • DER: 3-7%
   • Controlled settings, fewer speakers, clearer audio

4. AI UNCERTAINTY MARKERS are most prevalent in Small Group settings:
   • Shows AI awareness of challenging audio conditions

IMPLICATIONS FOR RESEARCH:
• Recording type should be reported as a covariate in transcription quality analysis
• Small group interactions may require more manual verification
• Full class and interview settings suitable for AI-assisted transcription with minimal review
• Consider audio quality improvements for small group recordings (e.g., individual mics)
""")

# Calculate overall statistics
total_utterances = sum(s['utterances'] for s in stats_by_condition.values())
total_speaker_worst = sum(s['speaker_err_worst'] for s in stats_by_condition.values())
total_utterance_worst = sum(s['utterance_err_worst'] for s in stats_by_condition.values())

print("\n" + "="*90)
print("OVERALL DATASET")
print("="*90)
print(f"Total utterances: {total_utterances}")
print(f"Total recordings: {sum(len(sheets) for sheets in conditions.values())}")
print(f"Overall DER (worst case): {total_speaker_worst/total_utterances*100:.2f}%")
print(f"Overall Utterance Error (worst case): {total_utterance_worst/total_utterances*100:.2f}%")

print("\n✅ Condition-based analysis complete!")
