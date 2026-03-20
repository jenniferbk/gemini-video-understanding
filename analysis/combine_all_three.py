import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from copy import copy

# File paths
jen_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_jen.xlsx'
annaB_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_annaB.xlsx'
uyi_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_Uyi.xlsx'
output_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_all_combined.xlsx'

print("=== COMBINING ALL THREE FILES ===\n")
print("Structure:")
print("  Cols A-B: Transcript (shared)")
print("  Cols C-E: Jennifer's observations")
print("  Col  F:   Spacer")
print("  Cols G-I: Anna's observations")
print("  Col  J:   Spacer")
print("  Cols K-M: Uyi's observations")
print()

# Load workbooks
wb_jen = openpyxl.load_workbook(jen_path)
wb_anna = openpyxl.load_workbook(annaB_path)
wb_uyi = openpyxl.load_workbook(uyi_path)

# Create new workbook for output
wb_output = openpyxl.Workbook()
wb_output.remove(wb_output.active)  # Remove default sheet

# Get all sheets (use Jennifer's as the master list)
sheets_to_process = wb_jen.sheetnames

print(f"Processing {len(sheets_to_process)} sheets...\n")

for sheet_name in sheets_to_process:
    print(f"Processing: {sheet_name}")

    # Skip if sheet doesn't exist in all files
    if sheet_name not in wb_anna.sheetnames or sheet_name not in wb_uyi.sheetnames:
        print(f"  WARNING: {sheet_name} not in all files, skipping")
        continue

    ws_jen = wb_jen[sheet_name]
    ws_anna = wb_anna[sheet_name]
    ws_uyi = wb_uyi[sheet_name]

    # Create new sheet in output workbook
    ws_output = wb_output.create_sheet(sheet_name)

    # Get max rows
    max_row = max(ws_jen.max_row, ws_anna.max_row, ws_uyi.max_row)

    # Process each row
    for row_idx in range(1, max_row + 1):
        # Columns A-B: Transcript (from Jennifer's file, same in all)
        for col_idx in [1, 2]:  # A, B
            cell_jen = ws_jen.cell(row=row_idx, column=col_idx)
            cell_output = ws_output.cell(row=row_idx, column=col_idx)

            if cell_jen.value is not None:
                cell_output.value = cell_jen.value

            if cell_jen.has_style:
                cell_output.font = copy(cell_jen.font)
                cell_output.border = copy(cell_jen.border)
                cell_output.fill = copy(cell_jen.fill)
                cell_output.number_format = copy(cell_jen.number_format)
                cell_output.protection = copy(cell_jen.protection)
                cell_output.alignment = copy(cell_jen.alignment)

        # Columns C-E: Jennifer's observations (from jen cols C-E)
        for col_idx in [3, 4, 5]:  # C, D, E
            cell_jen = ws_jen.cell(row=row_idx, column=col_idx)
            cell_output = ws_output.cell(row=row_idx, column=col_idx)

            if cell_jen.value is not None:
                cell_output.value = cell_jen.value

            if cell_jen.has_style:
                cell_output.font = copy(cell_jen.font)
                cell_output.border = copy(cell_jen.border)
                cell_output.fill = copy(cell_jen.fill)
                cell_output.number_format = copy(cell_jen.number_format)
                cell_output.protection = copy(cell_jen.protection)
                cell_output.alignment = copy(cell_jen.alignment)

        # Column F: Spacer
        ws_output.cell(row=row_idx, column=6).value = None

        # Columns G-I: Anna's observations (from anna cols C-E)
        for i, col_idx in enumerate([7, 8, 9]):  # G, H, I
            source_col = i + 3  # C, D, E (3, 4, 5)
            cell_anna = ws_anna.cell(row=row_idx, column=source_col)
            cell_output = ws_output.cell(row=row_idx, column=col_idx)

            if cell_anna.value is not None:
                cell_output.value = cell_anna.value

            if cell_anna.has_style:
                cell_output.font = copy(cell_anna.font)
                cell_output.border = copy(cell_anna.border)
                cell_output.fill = copy(cell_anna.fill)
                cell_output.number_format = copy(cell_anna.number_format)
                cell_output.protection = copy(cell_anna.protection)
                cell_output.alignment = copy(cell_anna.alignment)

        # Column J: Spacer
        ws_output.cell(row=row_idx, column=10).value = None

        # Columns K-M: Uyi's observations (from uyi cols C-E)
        for i, col_idx in enumerate([11, 12, 13]):  # K, L, M
            source_col = i + 3  # C, D, E (3, 4, 5)
            cell_uyi = ws_uyi.cell(row=row_idx, column=source_col)
            cell_output = ws_output.cell(row=row_idx, column=col_idx)

            if cell_uyi.value is not None:
                cell_output.value = cell_uyi.value

            if cell_uyi.has_style:
                cell_output.font = copy(cell_uyi.font)
                cell_output.border = copy(cell_uyi.border)
                cell_output.fill = copy(cell_uyi.fill)
                cell_output.number_format = copy(cell_uyi.number_format)
                cell_output.protection = copy(cell_uyi.protection)
                cell_output.alignment = copy(cell_uyi.alignment)

    # Add name labels in row 20
    # Jennifer label (spans C-E)
    cell_jen_label = ws_output.cell(row=20, column=3)
    cell_jen_label.value = "Jennifer"
    cell_jen_label.font = Font(bold=True, size=12)
    cell_jen_label.alignment = Alignment(horizontal='center')

    # Anna label (spans G-I)
    cell_anna_label = ws_output.cell(row=20, column=7)
    cell_anna_label.value = "Anna"
    cell_anna_label.font = Font(bold=True, size=12)
    cell_anna_label.alignment = Alignment(horizontal='center')

    # Uyi label (spans K-M)
    cell_uyi_label = ws_output.cell(row=20, column=11)
    cell_uyi_label.value = "Uyi"
    cell_uyi_label.font = Font(bold=True, size=12)
    cell_uyi_label.alignment = Alignment(horizontal='center')

    # Copy column widths
    for col_idx in [1, 2]:
        ws_output.column_dimensions[get_column_letter(col_idx)].width = \
            ws_jen.column_dimensions[get_column_letter(col_idx)].width

    # Jennifer's columns
    for col_idx in [3, 4, 5]:
        ws_output.column_dimensions[get_column_letter(col_idx)].width = \
            ws_jen.column_dimensions[get_column_letter(col_idx)].width

    # Spacer
    ws_output.column_dimensions['F'].width = 2

    # Anna's columns
    for i, col_idx in enumerate([7, 8, 9]):
        source_col = get_column_letter(i + 3)
        ws_output.column_dimensions[get_column_letter(col_idx)].width = \
            ws_anna.column_dimensions[source_col].width

    # Spacer
    ws_output.column_dimensions['J'].width = 2

    # Uyi's columns
    for i, col_idx in enumerate([11, 12, 13]):
        source_col = get_column_letter(i + 3)
        ws_output.column_dimensions[get_column_letter(col_idx)].width = \
            ws_uyi.column_dimensions[source_col].width

    print(f"  ✓ Completed ({max_row} rows)")

# Save output
print(f"\nSaving to: {output_path}")
wb_output.save(output_path)

wb_jen.close()
wb_anna.close()
wb_uyi.close()
wb_output.close()

print("\n✅ DONE! All three files combined successfully.")
print(f"\nNew file created: transcripts_all_combined.xlsx")
print(f"Total sheets: {len(sheets_to_process)}")
print("\nStructure:")
print("  - Row 20: Colleague names (Jennifer, Anna, Uyi)")
print("  - Cols A-B: Transcript")
print("  - Cols C-E: Jennifer's observations")
print("  - Cols G-I: Anna's observations")
print("  - Cols K-M: Uyi's observations")
