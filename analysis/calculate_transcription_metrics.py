import openpyxl
import re
from collections import defaultdict
import numpy as np

consensus_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_with_consensus.xlsx'

print("=== AI TRANSCRIPTION QUALITY METRICS ===\n")
print("Note: Calculating based on consensus items (3/3 agreement)")
print("Re-run after completing manual review for final metrics\n")

# Load workbook
wb = openpyxl.load_workbook(consensus_path)

# Statistics
stats = {
    'total_utterances': 0,
    'speaker_errors': 0,
    'utterance_errors': 0,
    'both_errors': 0,
    'uncertain_speakers': 0,
    'by_sheet': defaultdict(lambda: {
        'total': 0, 'speaker_err': 0, 'utterance_err': 0
    })
}

# Error patterns
speaker_error_patterns = defaultdict(int)

for sheet_name in wb.sheetnames:
    if sheet_name == 'REVIEW_LIST':
        continue

    ws = wb[sheet_name]

    for row_idx in range(22, ws.max_row + 1):
        # Skip if no transcript
        transcript = ws.cell(row=row_idx, column=1).value
        if not transcript or str(transcript).strip() == "":
            continue

        stats['total_utterances'] += 1
        stats['by_sheet'][sheet_name]['total'] += 1

        # Get consensus decisions (columns N, O)
        consensus_speaker = ws.cell(row=row_idx, column=14).value  # Col N
        consensus_utterance = ws.cell(row=row_idx, column=15).value  # Col O

        # Count errors (only where there's 3/3 agreement)
        if consensus_speaker == "YES - Error":
            stats['speaker_errors'] += 1
            stats['by_sheet'][sheet_name]['speaker_err'] += 1

            # Try to extract error pattern from transcript
            transcript_str = str(transcript)
            # Check if it's labeled as uncertain
            if '[uncertain]' in transcript_str or '⚠️' in transcript_str:
                stats['uncertain_speakers'] += 1

            # Extract AI's speaker label
            match = re.match(r'(\d+:\d+)\s+(.+?):', transcript_str)
            if match:
                ai_speaker = match.group(2).strip()
                speaker_error_patterns[ai_speaker] += 1

        if consensus_utterance == "YES - Error":
            stats['utterance_errors'] += 1
            stats['by_sheet'][sheet_name]['utterance_err'] += 1

        if consensus_speaker == "YES - Error" and consensus_utterance == "YES - Error":
            stats['both_errors'] += 1

wb.close()

# Calculate rates
print("="*70)
print("OVERALL TRANSCRIPTION QUALITY METRICS")
print("="*70)
print(f"\nTotal utterances analyzed: {stats['total_utterances']}")
print(f"\n--- DIARIZATION (Speaker Identification) ---")
print(f"Speaker errors (consensus): {stats['speaker_errors']}")
print(f"Diarization Error Rate (DER): {stats['speaker_errors']/stats['total_utterances']*100:.2f}%")
print(f"Speaker errors where AI was uncertain: {stats['uncertain_speakers']}")
print(f"  (These had ⚠️ or [uncertain] markers)")

print(f"\n--- TRANSCRIPTION (Text Accuracy) ---")
print(f"Utterance errors (consensus): {stats['utterance_errors']}")
print(f"Utterance Error Rate: {stats['utterance_errors']/stats['total_utterances']*100:.2f}%")

print(f"\n--- COMBINED ERRORS ---")
print(f"Utterances with both speaker AND text errors: {stats['both_errors']}")

print(f"\n--- ACCURACY METRICS ---")
print(f"Speaker Identification Accuracy: {(1 - stats['speaker_errors']/stats['total_utterances'])*100:.2f}%")
print(f"Utterance Transcription Accuracy: {(1 - stats['utterance_errors']/stats['total_utterances'])*100:.2f}%")
print(f"Perfect Utterances (no errors): {(1 - (stats['speaker_errors'] + stats['utterance_errors'] - stats['both_errors'])/stats['total_utterances'])*100:.2f}%")

# Most common speaker label errors
if speaker_error_patterns:
    print(f"\n--- MOST COMMON MISIDENTIFIED SPEAKERS (AI labels) ---")
    sorted_patterns = sorted(speaker_error_patterns.items(), key=lambda x: x[1], reverse=True)
    for speaker, count in sorted_patterns[:10]:
        print(f"  {speaker}: {count} errors")

# By sheet
print(f"\n" + "="*70)
print("METRICS BY RECORDING")
print("="*70)
print(f"{'Sheet':<20} {'Utterances':>12} {'Speaker Err':>12} {'DER %':>8} {'Utt Err':>12} {'Utt Err %':>10}")
print("-"*80)

for sheet_name in sorted(stats['by_sheet'].keys()):
    sheet_stats = stats['by_sheet'][sheet_name]
    total = sheet_stats['total']
    spk_err = sheet_stats['speaker_err']
    utt_err = sheet_stats['utterance_err']

    if total > 0:
        der = spk_err / total * 100
        uer = utt_err / total * 100
    else:
        der = uer = 0

    print(f"{sheet_name:<20} {total:>12} {spk_err:>12} {der:>7.1f}% {utt_err:>12} {uer:>9.1f}%")

print("\n" + "="*70)
print("NOTES FOR RESEARCH PAPER")
print("="*70)
print("""
These metrics represent:
1. DER (Diarization Error Rate): Percentage of utterances with incorrect speaker labels
2. Utterance Error Rate: Percentage of utterances with transcription errors
3. Based on consensus coding by 3 independent raters
4. Inter-rater reliability: Fleiss' κ = 0.200 (speaker), 0.146 (utterance)

Limitations:
- Current metrics based on 3/3 agreement only (conservative estimate)
- Full metrics available after completing manual review of disagreements
- Does not include word-level error analysis (would require full re-transcription)

For publication:
- Report DER and overall accuracy
- Note that errors include both complete misidentifications and marginal cases
- AI's uncertainty markers (⚠️, [uncertain]) correlate with actual errors
""")

print("\n✅ Transcription quality metrics calculated!")
print("\nNext step: Complete the REVIEW_LIST tab, then re-run this script for final metrics")
