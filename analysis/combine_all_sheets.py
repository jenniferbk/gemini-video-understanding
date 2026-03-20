import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

# File paths
annaB_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_annaB.xlsx'
uyi_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_Uyi.xlsx'
output_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_combined_complete.xlsx'

print("=== COMBINING ALL SHEETS ===\n")

# Load workbooks
wb_anna = openpyxl.load_workbook(annaB_path)
wb_uyi = openpyxl.load_workbook(uyi_path)

# Create new workbook for output
wb_output = openpyxl.Workbook()
wb_output.remove(wb_output.active)  # Remove default sheet

# Get all sheets from Uyi (skip Sheet1 from annaB)
sheets_to_process = [sheet for sheet in wb_uyi.sheetnames]

print(f"Processing {len(sheets_to_process)} sheets...\n")

for sheet_name in sheets_to_process:
    print(f"Processing: {sheet_name}")

    # Skip if sheet doesn't exist in both files
    if sheet_name not in wb_anna.sheetnames:
        print(f"  WARNING: {sheet_name} not in annaB file, skipping")
        continue

    ws_anna = wb_anna[sheet_name]
    ws_uyi = wb_uyi[sheet_name]

    # Create new sheet in output workbook
    ws_output = wb_output.create_sheet(sheet_name)

    # Get dimensions
    max_row = max(ws_anna.max_row, ws_uyi.max_row)

    # Process each row
    for row_idx in range(1, max_row + 1):
        # Columns A-B: Copy from annaB (same as Uyi)
        for col_idx in [1, 2]:  # A, B
            cell_anna = ws_anna.cell(row=row_idx, column=col_idx)
            cell_output = ws_output.cell(row=row_idx, column=col_idx)

            # Copy value or formula
            if cell_anna.value is not None:
                if str(cell_anna.value).startswith('='):
                    cell_output.value = cell_anna.value
                else:
                    cell_output.value = cell_anna.value

            # Copy formatting
            if cell_anna.has_style:
                cell_output.font = copy(cell_anna.font)
                cell_output.border = copy(cell_anna.border)
                cell_output.fill = copy(cell_anna.fill)
                cell_output.number_format = copy(cell_anna.number_format)
                cell_output.protection = copy(cell_anna.protection)
                cell_output.alignment = copy(cell_anna.alignment)

        # Columns C-E: Anna's observations (from annaB cols C-E)
        for col_idx in [3, 4, 5]:  # C, D, E
            cell_anna = ws_anna.cell(row=row_idx, column=col_idx)
            cell_output = ws_output.cell(row=row_idx, column=col_idx)

            if cell_anna.value is not None:
                if str(cell_anna.value).startswith('='):
                    cell_output.value = cell_anna.value
                else:
                    cell_output.value = cell_anna.value

            if cell_anna.has_style:
                cell_output.font = copy(cell_anna.font)
                cell_output.border = copy(cell_anna.border)
                cell_output.fill = copy(cell_anna.fill)
                cell_output.number_format = copy(cell_anna.number_format)
                cell_output.protection = copy(cell_anna.protection)
                cell_output.alignment = copy(cell_anna.alignment)

        # Columns F-H: Jennifer's observations (leave empty for now)
        for col_idx in [6, 7, 8]:  # F, G, H
            cell_output = ws_output.cell(row=row_idx, column=col_idx)
            cell_output.value = None

        # Columns I-K: Uyi's observations (from Uyi cols C-E)
        for i, col_idx in enumerate([9, 10, 11]):  # I, J, K
            source_col = i + 3  # C, D, E (3, 4, 5)
            cell_uyi = ws_uyi.cell(row=row_idx, column=source_col)
            cell_output = ws_output.cell(row=row_idx, column=col_idx)

            if cell_uyi.value is not None:
                if str(cell_uyi.value).startswith('='):
                    cell_output.value = cell_uyi.value
                else:
                    cell_output.value = cell_uyi.value

            if cell_uyi.has_style:
                cell_output.font = copy(cell_uyi.font)
                cell_output.border = copy(cell_uyi.border)
                cell_output.fill = copy(cell_uyi.fill)
                cell_output.number_format = copy(cell_uyi.number_format)
                cell_output.protection = copy(cell_uyi.protection)
                cell_output.alignment = copy(cell_uyi.alignment)

        # Column L: Empty
        cell_output = ws_output.cell(row=row_idx, column=12)
        cell_output.value = None

    # Copy column widths
    for col_idx in [1, 2]:
        ws_output.column_dimensions[get_column_letter(col_idx)].width = \
            ws_anna.column_dimensions[get_column_letter(col_idx)].width

    for col_idx in [3, 4, 5]:
        ws_output.column_dimensions[get_column_letter(col_idx)].width = \
            ws_anna.column_dimensions[get_column_letter(col_idx)].width

    for col_idx in [9, 10, 11]:
        source_col = get_column_letter(col_idx - 6)  # Map to C, D, E
        target_col = get_column_letter(col_idx)
        ws_output.column_dimensions[target_col].width = \
            ws_uyi.column_dimensions[source_col].width

    print(f"  ✓ Completed ({max_row} rows)")

# Save output
print(f"\nSaving to: {output_path}")
wb_output.save(output_path)

wb_anna.close()
wb_uyi.close()
wb_output.close()

print("\n✅ DONE! All sheets combined successfully.")
print(f"\nNew file created: transcripts_combined_complete.xlsx")
print(f"Total sheets: {len(wb_output.sheetnames)}")
