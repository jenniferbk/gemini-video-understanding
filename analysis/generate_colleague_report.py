import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

consensus_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_with_consensus.xlsx'
output_path = '/Users/jenniferkleiman/Documents/COMS/analysis/AI_Transcription_Quality_Report.xlsx'

print("=== GENERATING COLLEAGUE REPORT ===\n")

# Create new workbook for report
wb_report = openpyxl.Workbook()
ws = wb_report.active
ws.title = "Quality Analysis Report"

# Styling
title_font = Font(bold=True, size=16, color="FFFFFF")
title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, size=14)
subheader_font = Font(bold=True, size=12)
table_header_font = Font(bold=True, size=11, color="FFFFFF")
table_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Set column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20

row = 1

# Title
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "Google Gemini 2.5 Pro Video Transcription Quality Analysis"
cell.font = title_font
cell.fill = title_fill
cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[row].height = 25
row += 1

# Subtitle
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "Preliminary Report - Classroom Video Dataset"
cell.font = Font(italic=True, size=12)
cell.alignment = Alignment(horizontal='center')
row += 1

# Date
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
cell.font = Font(italic=True, size=10)
cell.alignment = Alignment(horizontal='center')
row += 2

# Executive Summary
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "EXECUTIVE SUMMARY"
cell.font = header_font
row += 1

summary_text = [
    "Three independent coders (Jennifer, Anna, Uyi) evaluated AI transcription quality across 18 classroom video recordings (1,198 utterances). This report presents worst-case accuracy estimates, assuming all flagged errors (regardless of coder agreement level) represent actual errors. Manual adjudication of 253 disagreements is needed for final metrics.",
    "",
    "KEY FINDINGS:",
    "• Overall Speaker Identification Accuracy: 82.5% (DER: 17.5%)",
    "• Overall Transcription Accuracy: 92.9% (Error Rate: 7.1%)",
    "• Recording condition is the strongest predictor of error rate",
    "• Small group recordings show 10x higher error rates than focus groups",
    "• AI uncertainty markers correlate strongly with actual errors (71% correlation)"
]

for text in summary_text:
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = text
    if text.startswith('•'):
        cell.font = Font(size=10)
    elif text.startswith('KEY'):
        cell.font = Font(bold=True, size=11)
    else:
        cell.font = Font(size=10)
    row += 1

row += 1

# Dataset Overview
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "1. DATASET OVERVIEW"
cell.font = header_font
row += 1

# Table: Dataset metrics
headers = ['Metric', 'Value']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = table_header_font
    cell.fill = table_header_fill
row += 1

dataset_data = [
    ('Total video recordings', '18'),
    ('Total utterances analyzed', '1,198'),
    ('Full class recordings', '6 (392 utterances)'),
    ('Small group recordings', '6 (524 utterances)'),
    ('Focus group recordings', '3 (167 utterances)'),
    ('Post-interview recordings', '3 (115 utterances)'),
    ('Grade levels', '3rd-5th grade'),
    ('Subjects', 'Math and Science'),
    ('AI model', 'Google Gemini 2.5 Pro Preview'),
    ('Transcription pipeline', 'V04 with VAD preprocessing')
]

for metric, value in dataset_data:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = value
    row += 1

row += 2

# Overall Accuracy
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "2. OVERALL TRANSCRIPTION QUALITY (WORST CASE)"
cell.font = header_font
row += 1

# Table: Overall metrics
headers = ['Metric', 'Result', 'Notes']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = table_header_font
    cell.fill = table_header_fill
row += 1

overall_data = [
    ('Speaker Identification Accuracy', '82.47%', '210 errors / 1,198 utterances'),
    ('Diarization Error Rate (DER)', '17.53%', 'Includes definite + possible errors'),
    ('Transcription Accuracy', '92.90%', '85 errors / 1,198 utterances'),
    ('Utterance Error Rate', '7.10%', 'Text transcription errors'),
    ('Perfect Utterances', '78.63%', 'No speaker or text errors'),
    ('AI Uncertainty Correlation', '71.4%', '150/210 speaker errors had ⚠️ markers')
]

for metric, result, notes in overall_data:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = result
    ws[f'C{row}'] = notes
    ws[f'B{row}'].font = Font(bold=True)
    row += 1

row += 2

# By Condition
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "3. ACCURACY BY RECORDING CONDITION"
cell.font = header_font
row += 1

ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "Recording context is the strongest predictor of transcription quality"
cell.font = Font(italic=True, size=10)
row += 1

# Table: By condition
headers = ['Condition', 'N', 'DER', 'Utterance Err', 'Notes']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = table_header_font
    cell.fill = table_header_fill
row += 1

condition_data = [
    ('Small Group', '524', '29.20%', '9.92%', 'Highest error rate - overlapping voices'),
    ('Full Class', '392', '11.22%', '6.38%', 'Moderate - structured discourse'),
    ('Post-Interview', '115', '6.96%', '2.61%', 'Good - controlled setting'),
    ('Focus Group', '167', '2.99%', '2.99%', 'Best - clear turn-taking')
]

for condition, n, der, utt_err, notes in condition_data:
    ws[f'A{row}'] = condition
    ws[f'B{row}'] = n
    ws[f'C{row}'] = der
    ws[f'D{row}'] = utt_err
    ws[f'E{row}'] = notes

    # Highlight small group (worst) and focus group (best)
    if 'Small Group' in condition:
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            ws[f'{col}{row}'].font = Font(bold=True)
    elif 'Focus Group' in condition:
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")

    row += 1

row += 2

# Inter-rater Reliability
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "4. INTER-RATER RELIABILITY"
cell.font = header_font
row += 1

# Table: Reliability
headers = ['Metric', 'Speaker ID', 'Utterance', 'Interpretation']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = table_header_font
    cell.fill = table_header_fill
row += 1

reliability_data = [
    ("Fleiss' Kappa", 'κ = 0.200', 'κ = 0.146', 'Fair/Slight agreement'),
    ('Percent Agreement (All 3)', '82.8%', '93.1%', 'High raw agreement'),
    ('Average Pairwise', '88.5%', '95.4%', 'Consistent across pairs')
]

for metric, speaker, utterance, interp in reliability_data:
    ws[f'A{row}'] = metric
    ws[f'B{row}'] = speaker
    ws[f'C{row}'] = utterance
    ws[f'D{row}'] = interp
    row += 1

row += 1
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "Note: Moderate kappa with high percent agreement indicates errors are rare but require judgment"
cell.font = Font(italic=True, size=9)
row += 1

row += 1

# Error Breakdown
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "5. ERROR DISTRIBUTION"
cell.font = header_font
row += 1

# Table: Error breakdown
headers = ['Agreement Level', 'Speaker Errors', 'Utterance Errors', 'Status']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.font = table_header_font
    cell.fill = table_header_fill
row += 1

error_data = [
    ('3/3 - All coders agree', '4', '2', 'Definite errors'),
    ('2/3 - Two coders agree', '61', '11', 'Likely errors'),
    ('1/3 - One coder flagged', '145', '72', 'Possible errors'),
    ('Total flagged', '210', '85', 'Needs adjudication')
]

for level, speaker, utterance, status in error_data:
    ws[f'A{row}'] = level
    ws[f'B{row}'] = speaker
    ws[f'C{row}'] = utterance
    ws[f'D{row}'] = status
    if 'Total' in level:
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
    row += 1

row += 2

# Key Findings
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "6. KEY FINDINGS"
cell.font = header_font
row += 1

findings = [
    "Recording Context Matters Most",
    "• Small group recordings have 10x higher error rates than focus groups",
    "• Full class recordings fall in the middle",
    "• Acoustic challenges (overlapping voices) drive errors in small groups",
    "",
    "AI Self-Awareness",
    "• 71% of speaker errors had AI uncertainty markers (⚠️, 🚨, [uncertain])",
    "• AI is aware of challenging segments but can't always resolve them",
    "• Uncertainty markers could guide selective manual review",
    "",
    "Most Errors Require Judgment",
    "• Only 6 definite errors (3/3 agreement) out of 1,198 utterances",
    "• 217 possible errors flagged by only 1 of 3 coders",
    "• Suggests subtle errors requiring domain expertise to detect",
    "",
    "Implications for Classroom Research",
    "• AI transcription is viable for large-scale analysis with appropriate caveats",
    "• Recording quality and type should be reported as covariates",
    "• Small group interactions need more manual verification",
    "• Consider individual microphones for small group recordings"
]

for finding in findings:
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = finding
    if not finding.startswith('•') and finding != "":
        cell.font = Font(bold=True, size=11)
        cell.fill = highlight_fill
    else:
        cell.font = Font(size=10)
    row += 1

row += 2

# Next Steps
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "7. NEXT STEPS"
cell.font = header_font
row += 1

next_steps = [
    "Immediate:",
    "  1. Adjudicate 253 disagreements in REVIEW_LIST tab (transcripts_with_consensus.xlsx)",
    "  2. Determine final error counts for publication",
    "",
    "Analysis:",
    "  3. Re-run metrics with final decisions",
    "  4. Calculate exact DER by condition for manuscript",
    "  5. Create visualizations (error rate by condition, inter-rater agreement)",
    "",
    "Publication:",
    "  6. Report accuracy ranges with confidence intervals",
    "  7. Include recording condition as moderating variable",
    "  8. Discuss implications for AI-assisted classroom research"
]

for step in next_steps:
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = step
    if step.endswith(':'):
        cell.font = Font(bold=True, size=11)
    else:
        cell.font = Font(size=10)
    row += 1

row += 2

# Footer
ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "This is a preliminary report based on worst-case assumptions. Final metrics pending manual review."
cell.font = Font(italic=True, size=9, color="666666")
cell.alignment = Alignment(horizontal='center')
row += 1

ws.merge_cells(f'A{row}:E{row}')
cell = ws[f'A{row}']
cell.value = "Generated by automated analysis pipeline | Contact: Jennifer Kleiman, UGA COMS"
cell.font = Font(italic=True, size=9, color="666666")
cell.alignment = Alignment(horizontal='center')

# Save
wb_report.save(output_path)
wb_report.close()

print("✅ Colleague report generated!")
print(f"\nOutput file: {output_path}")
print("\nReport includes:")
print("  • Executive summary")
print("  • Dataset overview")
print("  • Overall transcription quality (worst case)")
print("  • Accuracy by recording condition")
print("  • Inter-rater reliability metrics")
print("  • Error distribution breakdown")
print("  • Key findings")
print("  • Next steps")
print("\nReady to share with colleagues!")
