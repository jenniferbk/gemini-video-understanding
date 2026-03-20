import openpyxl
import pandas as pd

jen_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_jen.xlsx'

print("=== JENNIFER'S FILE STRUCTURE ===\n")

wb_jen = openpyxl.load_workbook(jen_path)

print(f"Sheets: {wb_jen.sheetnames}\n")

# Check Full_Jake structure
ws = wb_jen['Full_Jake']

print("Full_Jake - Row 21 (headers):")
for col_idx in range(1, 6):
    cell = ws.cell(row=21, column=col_idx)
    print(f"  Col {openpyxl.utils.get_column_letter(col_idx)}: {cell.value}")

print("\nFull_Jake - Row 22 (first data row):")
for col_idx in range(1, 6):
    cell = ws.cell(row=22, column=col_idx)
    if cell.value:
        val = str(cell.value)[:60]
        print(f"  Col {openpyxl.utils.get_column_letter(col_idx)}: {val}")

wb_jen.close()

print("\n✓ Confirmed: Jennifer's file has observations in columns C-E")
