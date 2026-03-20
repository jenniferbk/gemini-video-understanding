import openpyxl

output_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_all_combined.xlsx'

wb = openpyxl.load_workbook(output_path)

print("=== FINAL VERIFICATION ===\n")

print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheets: {wb.sheetnames}\n")

# Verify Full_Jake sheet
sheet_name = 'Full_Jake'
ws = wb[sheet_name]

print(f"{'='*80}")
print(f"Sheet: {sheet_name}")
print(f"{'='*80}\n")

print("Row 20 (Name labels):")
for col_idx in range(1, 14):
    cell = ws.cell(row=20, column=col_idx)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"  {col_letter}: {cell.value}")

print("\nRow 21 (Headers):")
for col_idx in range(1, 14):
    cell = ws.cell(row=21, column=col_idx)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"  {col_letter}: {cell.value}")

print("\nRow 22 (First data row - showing non-empty cells):")
for col_idx in range(1, 14):
    cell = ws.cell(row=22, column=col_idx)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        val = str(cell.value)[:60]
        print(f"  {col_letter}: {val}")

print("\n" + "="*80)
print("COLUMN MAPPING:")
print("="*80)
print("  A-B:  Transcript (shared)")
print("  C-E:  Jennifer's observations")
print("  F:    Spacer")
print("  G-I:  Anna's observations")
print("  J:    Spacer")
print("  K-M:  Uyi's observations")

wb.close()

print("\n✅ Verification complete!")
print(f"File: transcripts_all_combined.xlsx")
