import openpyxl
import pandas as pd
from collections import defaultdict

combined_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_all_combined.xlsx'

print("=== ANALYZING CODING STRUCTURE ===\n")

wb = openpyxl.load_workbook(combined_path)
ws = wb['Full_Jake']

# Sample rows to understand what's being coded
print("Sample coding from Full_Jake (first 50 rows):\n")

coding_examples = {
    'jennifer': [],
    'anna': [],
    'uyi': []
}

for row_idx in range(22, 72):  # Sample 50 rows after header
    # Get transcript
    transcript = ws.cell(row=row_idx, column=1).value

    # Jennifer (C-E)
    jen_speaker = ws.cell(row=row_idx, column=3).value
    jen_utterance = ws.cell(row=row_idx, column=4).value
    jen_notes = ws.cell(row=row_idx, column=5).value

    # Anna (G-I)
    anna_speaker = ws.cell(row=row_idx, column=7).value
    anna_utterance = ws.cell(row=row_idx, column=8).value
    anna_notes = ws.cell(row=row_idx, column=9).value

    # Uyi (K-M)
    uyi_speaker = ws.cell(row=row_idx, column=11).value
    uyi_utterance = ws.cell(row=row_idx, column=12).value
    uyi_notes = ws.cell(row=row_idx, column=13).value

    # Collect examples where people coded
    if jen_speaker or jen_utterance or jen_notes:
        coding_examples['jennifer'].append({
            'row': row_idx,
            'transcript': str(transcript)[:60] if transcript else "",
            'speaker': jen_speaker,
            'utterance': jen_utterance,
            'notes': jen_notes
        })

    if anna_speaker or anna_utterance or anna_notes:
        coding_examples['anna'].append({
            'row': row_idx,
            'transcript': str(transcript)[:60] if transcript else "",
            'speaker': anna_speaker,
            'utterance': anna_utterance,
            'notes': anna_notes
        })

    if uyi_speaker or uyi_utterance or uyi_notes:
        coding_examples['uyi'].append({
            'row': row_idx,
            'transcript': str(transcript)[:60] if transcript else "",
            'speaker': uyi_speaker,
            'utterance': uyi_utterance,
            'notes': uyi_notes
        })

# Show examples
for coder, examples in coding_examples.items():
    print(f"\n{coder.upper()} - {len(examples)} coded rows in sample:")
    for ex in examples[:5]:  # Show first 5
        print(f"  Row {ex['row']}:")
        print(f"    Transcript: {ex['transcript']}")
        print(f"    Speaker ID: {ex['speaker']}")
        print(f"    Utterance:  {ex['utterance']}")
        print(f"    Notes:      {ex['notes']}")
    if len(examples) > 5:
        print(f"    ... and {len(examples) - 5} more")

wb.close()

print("\n\n=== NEXT STEPS ===")
print("1. What are you coding?")
print("   - Speaker ID corrections?")
print("   - Utterance corrections?")
print("   - Quality notes?")
print("\n2. For transcription metrics, I'll need:")
print("   - AI transcript (column A)")
print("   - Your consensus corrections")
print("   - Speaker labels from AI vs your corrections")
