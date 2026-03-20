import openpyxl

output_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_all_combined.xlsx'

wb = openpyxl.load_workbook(output_path)
ws = wb['Full_Jake']

print("=== COMPLETE DATA CHECK - Full_Jake ===\n")
print("Looking for first row with data in any observation column...\n")

for row_idx in range(22, 50):
    jen_data = []
    anna_data = []
    uyi_data = []

    # Jennifer (C-E)
    for col_idx in [3, 4, 5]:
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            jen_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}:{str(cell.value)[:30]}")

    # Anna (G-I)
    for col_idx in [7, 8, 9]:
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            anna_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}:{str(cell.value)[:30]}")

    # Uyi (K-M)
    for col_idx in [11, 12, 13]:
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            uyi_data.append(f"{openpyxl.utils.get_column_letter(col_idx)}:{str(cell.value)[:30]}")

    if jen_data or anna_data or uyi_data:
        print(f"Row {row_idx}:")
        if jen_data:
            print(f"  Jennifer: {', '.join(jen_data)}")
        if anna_data:
            print(f"  Anna:     {', '.join(anna_data)}")
        if uyi_data:
            print(f"  Uyi:      {', '.join(uyi_data)}")
        print()

wb.close()

print("\n✅ Data check complete!")
