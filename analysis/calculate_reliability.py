import openpyxl
import pandas as pd
import numpy as np
from sklearn.metrics import cohen_kappa_score
from collections import defaultdict

consensus_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_with_consensus.xlsx'

print("=== INTER-RATER RELIABILITY ANALYSIS ===\n")

# Load workbook
wb = openpyxl.load_workbook(consensus_path)

# Aggregate data across all sheets
all_data = {
    'speaker': {'jennifer': [], 'anna': [], 'uyi': []},
    'utterance': {'jennifer': [], 'anna': [], 'uyi': []}
}

total_rows = 0
total_sheets = 0

for sheet_name in wb.sheetnames:
    total_sheets += 1
    ws = wb[sheet_name]

    for row_idx in range(22, ws.max_row + 1):
        # Skip if no transcript
        transcript = ws.cell(row=row_idx, column=1).value
        if not transcript or str(transcript).strip() == "":
            continue

        total_rows += 1

        # Get coding (1 = error flagged, 0 = no error flagged)
        # Jennifer (C-E)
        jen_speaker = 1 if ws.cell(row=row_idx, column=3).value else 0
        jen_utterance = 1 if ws.cell(row=row_idx, column=4).value else 0

        # Anna (G-I)
        anna_speaker = 1 if ws.cell(row=row_idx, column=7).value else 0
        anna_utterance = 1 if ws.cell(row=row_idx, column=8).value else 0

        # Uyi (K-M)
        uyi_speaker = 1 if ws.cell(row=row_idx, column=11).value else 0
        uyi_utterance = 1 if ws.cell(row=row_idx, column=12).value else 0

        # Store
        all_data['speaker']['jennifer'].append(jen_speaker)
        all_data['speaker']['anna'].append(anna_speaker)
        all_data['speaker']['uyi'].append(uyi_speaker)

        all_data['utterance']['jennifer'].append(jen_utterance)
        all_data['utterance']['anna'].append(anna_utterance)
        all_data['utterance']['uyi'].append(uyi_utterance)

wb.close()

print(f"Total sheets analyzed: {total_sheets}")
print(f"Total transcript rows: {total_rows}\n")

# Calculate pairwise Cohen's Kappa
print("="*70)
print("PAIRWISE INTER-RATER RELIABILITY (Cohen's Kappa)")
print("="*70)

for category in ['speaker', 'utterance']:
    print(f"\n{category.upper()} ID ERRORS:\n")

    # Jennifer vs Anna
    kappa_ja = cohen_kappa_score(
        all_data[category]['jennifer'],
        all_data[category]['anna']
    )
    print(f"  Jennifer vs Anna:  κ = {kappa_ja:.3f}")

    # Jennifer vs Uyi
    kappa_ju = cohen_kappa_score(
        all_data[category]['jennifer'],
        all_data[category]['uyi']
    )
    print(f"  Jennifer vs Uyi:   κ = {kappa_ju:.3f}")

    # Anna vs Uyi
    kappa_au = cohen_kappa_score(
        all_data[category]['anna'],
        all_data[category]['uyi']
    )
    print(f"  Anna vs Uyi:       κ = {kappa_au:.3f}")

    # Average
    avg_kappa = (kappa_ja + kappa_ju + kappa_au) / 3
    print(f"  Average κ:         {avg_kappa:.3f}")

# Calculate Fleiss' Kappa (for 3 raters)
def fleiss_kappa(matrix):
    """
    Calculate Fleiss' kappa for multiple raters
    matrix: n_items x n_categories (here: n_rows x 2 [no error, error])
    """
    n_items, n_categories = matrix.shape
    n_raters = matrix.sum(axis=1)[0]  # Sum across categories for first item

    # Calculate p_j (proportion of all assignments in category j)
    p_j = matrix.sum(axis=0) / (n_items * n_raters)

    # Calculate P_i (extent of agreement for item i)
    P_i = (np.sum(matrix ** 2, axis=1) - n_raters) / (n_raters * (n_raters - 1))

    # Calculate P_bar (mean of P_i)
    P_bar = np.mean(P_i)

    # Calculate P_e_bar (expected agreement by chance)
    P_e_bar = np.sum(p_j ** 2)

    # Calculate kappa
    kappa = (P_bar - P_e_bar) / (1 - P_e_bar)

    return kappa, P_bar, P_e_bar

print("\n" + "="*70)
print("FLEISS' KAPPA (All 3 Raters)")
print("="*70)

for category in ['speaker', 'utterance']:
    # Build matrix: rows = items, columns = categories (0=no error, 1=error)
    # For each row, count how many raters said 0 vs 1
    matrix = []
    for i in range(total_rows):
        jen_val = all_data[category]['jennifer'][i]
        anna_val = all_data[category]['anna'][i]
        uyi_val = all_data[category]['uyi'][i]

        # Count votes for each category
        no_error_count = (jen_val == 0) + (anna_val == 0) + (uyi_val == 0)
        error_count = (jen_val == 1) + (anna_val == 1) + (uyi_val == 1)

        matrix.append([no_error_count, error_count])

    matrix = np.array(matrix)
    kappa, p_bar, p_e = fleiss_kappa(matrix)

    print(f"\n{category.upper()} ID ERRORS:")
    print(f"  Fleiss' κ = {kappa:.3f}")
    print(f"  Observed agreement (P̄) = {p_bar:.3f}")
    print(f"  Expected agreement (P̄ₑ) = {p_e:.3f}")

# Calculate simple percent agreement
print("\n" + "="*70)
print("PERCENT AGREEMENT")
print("="*70)

for category in ['speaker', 'utterance']:
    print(f"\n{category.upper()} ID ERRORS:")

    # Pairwise
    ja_agree = sum(all_data[category]['jennifer'][i] == all_data[category]['anna'][i]
                   for i in range(total_rows)) / total_rows
    ju_agree = sum(all_data[category]['jennifer'][i] == all_data[category]['uyi'][i]
                   for i in range(total_rows)) / total_rows
    au_agree = sum(all_data[category]['anna'][i] == all_data[category]['uyi'][i]
                   for i in range(total_rows)) / total_rows

    print(f"  Jennifer vs Anna:  {ja_agree*100:.1f}%")
    print(f"  Jennifer vs Uyi:   {ju_agree*100:.1f}%")
    print(f"  Anna vs Uyi:       {au_agree*100:.1f}%")
    print(f"  Average:           {(ja_agree + ju_agree + au_agree)/3*100:.1f}%")

    # All three agree
    all_agree = sum(
        all_data[category]['jennifer'][i] == all_data[category]['anna'][i] == all_data[category]['uyi'][i]
        for i in range(total_rows)
    ) / total_rows
    print(f"  All 3 agree:       {all_agree*100:.1f}%")

# Kappa interpretation guide
print("\n" + "="*70)
print("INTERPRETATION GUIDE (Landis & Koch, 1977)")
print("="*70)
print("  κ < 0.00:     Poor agreement")
print("  κ 0.00-0.20:  Slight agreement")
print("  κ 0.21-0.40:  Fair agreement")
print("  κ 0.41-0.60:  Moderate agreement")
print("  κ 0.61-0.80:  Substantial agreement")
print("  κ 0.81-1.00:  Almost perfect agreement")

print("\n✅ Inter-rater reliability analysis complete!")
