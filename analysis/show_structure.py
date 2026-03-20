import pandas as pd
import openpyxl

# Use openpyxl to see the raw Excel structure
combined_path = '/Users/jenniferkleiman/Documents/COMS/analysis/transcripts_combined.xlsx'

wb = openpyxl.load_workbook(combined_path)
ws = wb['Full_Jake']

print("=== RAW EXCEL STRUCTURE (First 3 rows, all columns) ===\n")

for row_idx in range(1, 4):  # First 3 rows
    print(f"Row {row_idx}:")
    for col_idx in range(1, 13):  # 12 columns
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val:
            val_str = str(val)[:50]
            print(f"  Col {col_idx:2d} (Excel col {openpyxl.utils.get_column_letter(col_idx)}): {val_str}")
        else:
            print(f"  Col {col_idx:2d} (Excel col {openpyxl.utils.get_column_letter(col_idx)}): [empty]")
    print()

print("\n=== ROW 21 (actual transcript data) ===\n")
row_idx = 21
for col_idx in range(1, 13):
    cell = ws.cell(row=row_idx, column=col_idx)
    val = cell.value
    if val:
        val_str = str(val)[:60]
        print(f"Col {openpyxl.utils.get_column_letter(col_idx):2s} (index {col_idx-1:2d}): {val_str}")

wb.close()

# Now let's see the pattern across Full_Jake, Full_Ava, and Full_Daisy
print("\n\n=== COMPARING THE 3 MANUALLY DONE SHEETS ===\n")

for sheet in ['Full_Jake', 'Full_Ava', 'Full_Daisy']:
    df = pd.read_excel(combined_path, sheet_name=sheet)
    print(f"{sheet}: {df.shape[1]} columns")
